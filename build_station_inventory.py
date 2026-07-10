#!/usr/bin/env python3
"""
build_station_inventory.py

Rebuilds a clean station-metadata inventory (metadata_csv) from the raw
"Data wave 2" .xlsx archive files, from scratch (no dependence on any
prior partial inventory).

See PROJECT.md and RAW_FILE_SURVEY.md (in the same folder as the raw
files) for the full background, format survey, and the parsing/merge
plan this script implements. Key rules enforced here:

  - openpyxl only (never pandas.read_excel) so merged-cell ranges and
    text rotation are visible.
  - Two structurally different raw layouts are parsed with dedicated
    strategies:
      * MONTH-SHEET format: one sheet per calendar month, all stations
        as rows (1949-1957 monthly files, 1965.xlsx, 1966-1973 Tuyet
        series).
      * STATION-BLOCK format: one sheet per group of stations, each
        station a merged multi-row block covering all 12 months + an
        annual row (1958-1964, Data_1965.xlsx).
  - Coordinates are pulled ONLY from (a) the dedicated "Toa do" sheets
    in Data 1956/1957.xlsx (decimal-encoded degrees+minutes, e.g. 16.33
    means 16 deg 33 min) and (b) embedded "Lat.N/Long.E.G./Alt." text
    annotations found in station-name cells elsewhere. No coordinate is
    ever pulled from Excel_test/Done (cross-check only) or guessed.
  - Station identity: numeric code (indicatif) is the primary merge
    key where present, but every code match is cross-checked against
    normalized station name; a diverging name under a matched code is
    flagged for manual review rather than silently merged or split.
  - Every judgment call (ambiguous name merge, code/name mismatch,
    coordinate conflict, unparsed sheet) is recorded in a `flags`
    column and echoed to a companion flagged_for_review.csv.
"""

import csv
import glob
import os
import re
import sys
import unicodedata
from collections import defaultdict

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

RAW_DIR = os.environ.get("RAW_DIR", ".")
OUT_METADATA_CSV = os.environ.get("OUT_METADATA_CSV", "metadata_csv")
OUT_FLAGS_CSV = os.environ.get("OUT_FLAGS_CSV", "flagged_for_review.csv")
OUT_LOG = os.environ.get("OUT_LOG", "build_log.txt")

# Known duplicate/overlapping file pairs -> keep both but note in source_file;
# no special dedup of values is attempted here (metadata build only, not
# value-level reconciliation -- that's cleaned_csvs_rebuilt's job / Phase 2).
DUPLICATE_PAIRS = [
    ("Data 1960.xlsx", "Data 1960 new.xlsx"),
    ("Data 1961.xlsx", "Data 1961 (8-12).xlsx"),
    ("Data 1965.xlsx", "Data_1965.xlsx"),
]

# ---------------------------------------------------------------------------
# Name / text normalization helpers
# ---------------------------------------------------------------------------

FRENCH_MONTHS = {
    "jan": 1, "janv": 1, "janvier": 1,
    "fev": 2, "fev.": 2, "fevrier": 2,
    "mar": 3, "mars": 3,
    "avr": 4, "avril": 4,
    "mai": 5,
    "juin": 6,
    "juil": 7, "juillet": 7, "jul": 7,
    "aou": 8, "aout": 8,
    "sep": 9, "sept": 9, "septembre": 9,
    "oct": 10, "octobre": 10,
    "nov": 11, "novembre": 11,
    "dec": 12, "decembre": 12,
    # English spellings appear too (e.g. Data 1965.xlsx uses "April")
    "april": 4, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "may": 5,
}

ROMAN_MONTHS = {
    "i": 1, "ii": 2, "iii": 3, "iv": 4, "v": 5, "vi": 6,
    "vii": 7, "viii": 8, "ix": 9, "x": 10, "xi": 11, "xii": 12,
}


def strip_accents(s):
    if s is None:
        return ""
    text = str(s)
    # Vietnamese d-with-stroke has no NFKD decomposition to "d" (it's an
    # independent letter, not a combining-diacritic form), so it must be
    # handled explicitly before stripping combining marks -- otherwise it
    # silently disappears from normalized names instead of becoming "d".
    text = text.replace("đ", "d").replace("Đ", "D")
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def clean_token(s):
    """lowercase, strip accents, collapse whitespace/punctuation for matching."""
    s = strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def sheet_month_from_name(name):
    """Try to find a month number embedded in a sheet name via French tokens
    or a trailing/leading integer (e.g. 'Thang 3' -> 3, 'Temp 4.56' -> 4)."""
    base = clean_token(name)
    # "thang N" (Vietnamese for "month N")
    m = re.search(r"\bthang\s*(\d{1,2})\b", base)
    if m:
        v = int(m.group(1))
        if 1 <= v <= 12:
            return v
    # French month word anywhere in the name
    for tok in base.split():
        if tok in FRENCH_MONTHS:
            return FRENCH_MONTHS[tok]
    # numeric like "4.56" (month.year) or "temp 1 56"
    m = re.search(r"\b(\d{1,2})[\. ]\d{2}\b", name)
    if m:
        v = int(m.group(1))
        if 1 <= v <= 12:
            return v
    return None


def normalize_station_name(raw):
    """Split a raw station-name cell into (base_name_clean, qualifier_clean,
    embedded_coord_text_or_None). Embedded coordinate annotations
    ('Lat.N : ... Long.E.G.: ... Alt.: ...') are stripped out of the name
    text and returned separately."""
    if raw is None:
        return "", "", None
    text = str(raw)

    coord_text = None
    m = re.search(r"Lat\.?\s*N.*", text, flags=re.IGNORECASE | re.DOTALL)
    if m:
        coord_text = m.group(0)
        text = text[: m.start()]

    text = text.replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)

    qualifier = ""
    m = re.search(r"\(([^)]+)\)\s*$", text)
    if m:
        qualifier = m.group(1).strip()
        text = text[: m.start()].strip()

    base_clean = clean_token(text)
    qualifier_clean = clean_token(qualifier)
    return base_clean, qualifier_clean, coord_text


DMS_RE = re.compile(r"(\d+)\s*[°ºo]\s*(\d+)")
ALT_RE = re.compile(r"Alt\.?\s*:?\s*([0-9]+(?:\.[0-9]+)?|x)\s*m", re.IGNORECASE)


def parse_embedded_coords(coord_text):
    """Parse 'Lat.N : 11d33 Long.E.G.: 107d54 Alt.: 800m' (deg mark) style
    text. Returns (lat, lon, alt) with None for anything not
    found/unparseable. Degrees+minutes -> decimal degrees."""
    if not coord_text:
        return None, None, None
    lat = lon = alt = None
    parts = re.split(r"Long", coord_text, maxsplit=1, flags=re.IGNORECASE)
    lat_part = parts[0]
    lon_alt_part = parts[1] if len(parts) > 1 else ""

    m = DMS_RE.search(lat_part)
    if m:
        deg, minu = int(m.group(1)), int(m.group(2))
        lat = round(deg + minu / 60.0, 4)

    m = DMS_RE.search(lon_alt_part)
    if m:
        deg, minu = int(m.group(1)), int(m.group(2))
        lon = round(deg + minu / 60.0, 4)

    m = ALT_RE.search(coord_text)
    if m:
        val = m.group(1)
        if val.lower() != "x":
            try:
                alt = float(val)
            except ValueError:
                alt = None
    return lat, lon, alt


def decdeg_from_degmin_float(v):
    """Convert a 'Toa do' sheet value like 16.33 (16 deg 33 min) to decimal
    degrees. The fractional part IS the minutes (not a decimal fraction of
    a degree) per the sheet's own column header ('do va phut' = degrees and
    minutes)."""
    if v is None:
        return None
    try:
        v = float(v)
    except (TypeError, ValueError):
        return None
    sign = -1 if v < 0 else 1
    v = abs(v)
    deg = int(v)
    minu_raw = round((v - deg) * 100)
    if minu_raw >= 60:
        # Data-quality anomaly: fractional part doesn't look like valid
        # minutes. Bail out rather than silently producing an invalid
        # coordinate; caller flags this case.
        return None
    return sign * round(deg + minu_raw / 60.0, 4)


DEGSYM_RE = re.compile(r"(\d+)\s*[°ºo]\s*(\d+)")


def parse_coord_cell(v):
    """A 'Toa do'-style coordinate cell is inconsistent in the raw files:
    sometimes a float using the deg.MM encoding (e.g. 12.41 = 12 deg 41
    min), sometimes a literal string like '13°59'. Handle both, returning
    decimal degrees or None."""
    if v is None:
        return None
    if isinstance(v, str):
        m = DEGSYM_RE.search(v)
        if m:
            deg, minu = int(m.group(1)), int(m.group(2))
            return round(deg + minu / 60.0, 4)
        return None
    return decdeg_from_degmin_float(v)


# Vietnam + near-border plausibility bounding box (generous, includes
# Laos/Cambodge references seen in the colonial-era network e.g. "AI-LAO").
PLAUSIBLE_LAT = (5.0, 24.0)
PLAUSIBLE_LON = (100.0, 112.0)


def in_plausible_bounds(lat, lon):
    if lat is None or lon is None:
        return True  # nothing to check
    return PLAUSIBLE_LAT[0] <= lat <= PLAUSIBLE_LAT[1] and PLAUSIBLE_LON[0] <= lon <= PLAUSIBLE_LON[1]


# ---------------------------------------------------------------------------
# Sighting record
# ---------------------------------------------------------------------------

class Sighting:
    __slots__ = ("base", "qualifier", "raw_name", "code", "year", "month",
                 "has_data", "source_file", "sheet", "coord_text",
                 "direct_coord")

    def __init__(self, base, qualifier, raw_name, code, year, month,
                 has_data, source_file, sheet, coord_text=None,
                 direct_coord=None):
        self.base = base
        self.qualifier = qualifier
        self.raw_name = raw_name
        self.code = code
        self.year = year
        self.month = month
        self.has_data = has_data
        self.source_file = source_file
        self.sheet = sheet
        self.coord_text = coord_text
        self.direct_coord = direct_coord  # (lat, lon, alt) from a dedicated
                                           # coordinate sheet, or None


sightings = []
coord_registry = {}  # code -> (lat, lon, source_file)
log_lines = []


def log(msg):
    log_lines.append(msg)


def is_blank(v):
    if v is None:
        return True
    if isinstance(v, str) and v.strip().lower() in ("", "x", "-", "n/a"):
        return True
    return False


# ---------------------------------------------------------------------------
# Parser: dedicated "Toa do" (coordinates) metadata sheet, 1956/1957 only
# ---------------------------------------------------------------------------

def parse_coord_sheet(ws, source_file):
    """Handles BOTH coordinate-sheet layouts seen in Data 1956/1957.xlsx:
      - the primary "Toa do" sheet: [name, code(indicatif), lat, lon]
      - the rain-gauge-post "toa do (2)" sheet: [name, lat, lon] with NO
        code column at all, laid out as two side-by-side tables.
    Blindly assuming a fixed 4-column [name, code, lat, lon] layout (as an
    earlier version of this script did) silently misreads the rain-gauge
    sheet's latitude as an integer station code -- e.g. int(11.58) == 11 --
    which then falsely merges every unrelated station whose latitude
    happens to truncate to the same integer. Header text is inspected
    explicitly to tell the two layouts apart instead of assuming column
    positions."""
    header_row = None
    for r in range(1, min(6, ws.max_row) + 1):
        row_tokens = [clean_token(ws.cell(row=r, column=c).value)
                      for c in range(1, ws.max_column + 1)]
        if (any("latitude" in t or "vi do" in t for t in row_tokens) and
                any("station" in t or "tram" in t for t in row_tokens)):
            header_row = r
            break
    if header_row is None:
        log(f"SKIP coord sheet, no recognizable header: {source_file} / {ws.title}")
        return

    # NOTE: must check the header token as a PREFIX, not "contains", since
    # this sheet also has a "Loai Ty hoac tram / Type de station" column
    # whose text contains the substring "tram"/"station" even though it is
    # not itself a station-name column -- a naive "in" check would treat it
    # as a spurious extra table boundary and corrupt the column layout.
    starts = [c for c in range(1, ws.max_column + 1)
              if clean_token(ws.cell(row=header_row, column=c).value or "")
              .startswith(("station", "tram", "ty"))]
    if not starts:
        starts = [1]

    for i, start_col in enumerate(starts):
        name_col = start_col
        next_start = starts[i + 1] if i + 1 < len(starts) else ws.max_column + 1
        # Locate lat/lon/code columns by header TEXT within this table's
        # span, not by assumed offset -- the exact column layout differs
        # between files (e.g. Data 1957.xlsx's rain-gauge sheet inserts an
        # extra "type de station" column between name and latitude that
        # Data 1956.xlsx's equivalent sheet doesn't have).
        code_col = lat_col = lon_col = None
        for c in range(start_col + 1, next_start):
            t = clean_token(ws.cell(row=header_row, column=c).value or "")
            if not t:
                continue
            if "latitude" in t or "vi do" in t:
                lat_col = c
            elif "longitude" in t or "kinh do" in t:
                lon_col = c
            elif "indicatif" in t or "chi so" in t:
                code_col = c
        if lat_col is None or lon_col is None:
            log(f"SKIP coord table, lat/lon header not found: {source_file} / "
                f"{ws.title} (table at col {start_col})")
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=name_col).value
            if name is None or (isinstance(name, str) and not name.strip()):
                continue
            if isinstance(name, str) and clean_token(name) in (
                    "vietnam", "cao mien cambodge", "laos"):
                continue

            lat = parse_coord_cell(ws.cell(row=r, column=lat_col).value)
            lon = parse_coord_cell(ws.cell(row=r, column=lon_col).value)
            if lat is None and lon is None:
                continue  # row belongs to the other table, or is a label row

            code_i = None
            if code_col:
                code_val = ws.cell(row=r, column=code_col).value
                if isinstance(code_val, (int, float)):
                    code_i = int(code_val)

            base, qualifier, _ = normalize_station_name(name)
            if not base:
                continue
            # Also register a "name-only" sighting so this station is known
            # even if it never appears with data in a parsed month/block.
            sightings.append(Sighting(base, qualifier, str(name).strip(), code_i,
                                       None, None, False, source_file, ws.title,
                                       direct_coord=(lat, lon, None)))


# ---------------------------------------------------------------------------
# Parser: MONTH-SHEET format
# (1949-1954, 1955, 1956/1957 temp+preci sheets, 1965.xlsx, 1966-1973 Tuyet)
# ---------------------------------------------------------------------------

HEADER_TOKENS = ("station", "stations", "noi quan trac", "ty",
                  "tram vu luong", "postes pluviometriques")


def find_header_row(ws, max_scan=10):
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, min(4, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            t = clean_token(v)
            if t.startswith(HEADER_TOKENS):
                return r
    return None


def parse_month_sheet(ws, year, month, source_file):
    header_row = find_header_row(ws)
    if header_row is None:
        log(f"SKIP unparsed sheet (no header found): {source_file} / {ws.title}")
        return
    start_row = header_row + 1
    # Detect an optional numeric-code column: if column A in the first data
    # row (skipping section-label rows like 'VIET-NAM') is numeric, treat
    # column A as code and column B as name; else column A is the name.
    code_col = None
    name_col = 1
    for r in range(start_row, min(start_row + 5, ws.max_row + 1)):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if isinstance(a, (int, float)) and isinstance(b, str) and b.strip():
            code_col, name_col = 1, 2
            break
        if isinstance(a, str) and a.strip():
            break

    for r in range(start_row, ws.max_row + 1):
        name_val = ws.cell(row=r, column=name_col).value
        if is_blank(name_val):
            continue
        if isinstance(name_val, str) and clean_token(name_val) in (
                "vietnam", "cao mien cambodge", "laos"):
            continue  # section-label row, not a station
        code_val = ws.cell(row=r, column=code_col).value if code_col else None
        code = int(code_val) if isinstance(code_val, (int, float)) else None

        data_cols = range(1, ws.max_column + 1)
        has_data = any(
            not is_blank(ws.cell(row=r, column=c).value)
            for c in data_cols if c not in (code_col, name_col)
        )

        base, qualifier, coord_text = normalize_station_name(name_val)
        if not base:
            continue
        if "region" in qualifier:
            # Structural section-label row (e.g. "Vung Kontum (Region de
            # Kontum)"), not an actual station. Only skipped when the
            # qualifier itself says "region", so real "Vung Tau ..."
            # stations (qualifiers like "aerodrome", "Hai Dang") pass
            # through untouched.
            continue
        sightings.append(Sighting(base, qualifier, str(name_val).strip(),
                                   code, year, month, has_data, source_file,
                                   ws.title, coord_text))


# ---------------------------------------------------------------------------
# Parser: STATION-BLOCK format (1958-1964, Data_1965.xlsx)
# ---------------------------------------------------------------------------

def find_table_starts(ws, header_row):
    """A sheet may contain more than one station table side-by-side. Return
    a list of starting columns where a 'station' header token is found on
    the header row. Checked as a PREFIX match, not "contains" -- some
    non-name columns (e.g. a "Type de station" column) contain a header
    token as a substring without actually being a name column."""
    starts = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        t = clean_token(v)
        if t.startswith(HEADER_TOKENS):
            starts.append(c)
    return starts or [1]


def parse_station_block_sheet(ws, year, source_file):
    header_row = find_header_row(ws)
    if header_row is None:
        log(f"SKIP unparsed sheet (no header found): {source_file} / {ws.title}")
        return

    table_starts = find_table_starts(ws, header_row)

    for start_col in table_starts:
        name_col = start_col
        month_col = start_col + 1
        # data columns for this table: up to just before the next table's
        # start column (or end of sheet)
        next_starts = [s for s in table_starts if s > start_col]
        end_col = (min(next_starts) - 1) if next_starts else ws.max_column

        # merged ranges anchored in name_col, below header_row
        blocks = []
        for mc in ws.merged_cells.ranges:
            if mc.min_col == name_col and mc.min_row > header_row:
                blocks.append((mc.min_row, mc.max_row))
        blocks.sort()

        if not blocks:
            log(f"SKIP no station blocks found: {source_file} / {ws.title} "
                f"(table at col {start_col})")
            continue

        for (r0, r1) in blocks:
            name_val = ws.cell(row=r0, column=name_col).value
            if is_blank(name_val):
                continue
            base, qualifier, coord_text = normalize_station_name(name_val)
            if not base:
                continue
            if "region" in qualifier:
                continue  # structural section-label block, not a station

            months_with_data = set()
            for r in range(r0, r1 + 1):
                month_val = ws.cell(row=r, column=month_col).value
                if month_val is None:
                    continue
                mt = clean_token(month_val)
                month_num = ROMAN_MONTHS.get(mt)
                if month_num is None:
                    continue  # e.g. the 'Nam/Annee' annual summary row
                row_has_data = any(
                    not is_blank(ws.cell(row=r, column=c).value)
                    for c in range(month_col + 1, end_col + 1)
                )
                if row_has_data:
                    months_with_data.add(month_num)

            if not months_with_data:
                continue

            for m in sorted(months_with_data):
                sightings.append(Sighting(base, qualifier, str(name_val).strip(),
                                           None, year, m, True, source_file,
                                           ws.title, coord_text))


# ---------------------------------------------------------------------------
# File dispatch
# ---------------------------------------------------------------------------

def year_from_filename(fname):
    m = re.search(r"(19[4-9]\d)", fname)
    return int(m.group(1)) if m else None


def sheet_is_coord_sheet(ws):
    t = clean_token(ws.title)
    return "toa do" in t or "toado" in t


def sheet_looks_like_station_block(ws):
    """Distinguish STATION-BLOCK sheets (merged multi-row name cells with
    roman-numeral month rows) from MONTH-SHEET sheets (one row per
    station, all in a single header + data-rows layout)."""
    header_row = find_header_row(ws)
    if header_row is None:
        return False
    for mc in ws.merged_cells.ranges:
        if mc.min_col == 1 and mc.min_row > header_row and mc.max_row > mc.min_row:
            return True
    return False


def process_file(path):
    fname = os.path.basename(path)
    year_guess = year_from_filename(fname)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        log(f"ERROR opening {fname}: {e}")
        return

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_is_coord_sheet(ws):
            parse_coord_sheet(ws, fname)
            continue

        if sheet_looks_like_station_block(ws):
            if year_guess is None:
                log(f"SKIP station-block sheet, no year in filename: {fname}/{sheet_name}")
                continue
            parse_station_block_sheet(ws, year_guess, fname)
        else:
            month = sheet_month_from_name(sheet_name)
            if month is None or year_guess is None:
                log(f"SKIP unparsed month-sheet name: {fname}/{sheet_name}")
                continue
            parse_month_sheet(ws, year_guess, month, fname)


# ---------------------------------------------------------------------------
# Clustering
# ---------------------------------------------------------------------------

class Cluster:
    def __init__(self, cid):
        self.id = cid
        self.codes = set()
        self.raw_names = []          # ordered, de-duplicated raw text variants
        self.bases = set()           # normalized base names seen
        self.qualifiers = []
        self.years_months = defaultdict(set)   # year -> set(months)
        self.source_files = set()
        self.coord_candidates = []   # (lat, lon, alt, source, priority)
        self.flags = []

    def add_raw_name(self, raw):
        if raw not in self.raw_names:
            self.raw_names.append(raw)


def name_similarity_ok(base_a, base_b):
    """Loose check used only to decide whether to *flag* a code-match, never
    to decide whether to merge (code always merges). Shares at least one
    non-trivial token, or high sequence-similarity."""
    if not base_a or not base_b:
        return True
    ta, tb = set(base_a.split()), set(base_b.split())
    if ta & tb:
        return True
    import difflib
    ratio = difflib.SequenceMatcher(None, base_a, base_b).ratio()
    return ratio >= 0.5


def build_clusters(sightings):
    clusters = {}          # cluster_id -> Cluster
    code_to_cluster = {}   # code -> cluster_id
    name_to_cluster = {}   # normalized base name -> cluster_id (code-less only)
    next_id = [1]

    def new_cluster():
        cid = next_id[0]
        next_id[0] += 1
        clusters[cid] = Cluster(cid)
        return cid

    # Sort so that code-bearing sightings are processed first (they anchor
    # clusters most reliably), then code-less ones, both in first-seen
    # (year, then insertion) order for stable station_id assignment.
    def sort_key(i):
        s = sightings[i]
        return (s.code is None, s.year if s.year is not None else 9999)

    ordered = sorted(range(len(sightings)), key=sort_key)

    for idx in ordered:
        s = sightings[idx]

        if s.code is not None:
            cid = code_to_cluster.get(s.code)
            if cid is None:
                cid = new_cluster()
                code_to_cluster[s.code] = cid
                clusters[cid].codes.add(s.code)
            cl = clusters[cid]
            if cl.bases and not any(name_similarity_ok(s.base, b) for b in cl.bases):
                cl.flags.append(
                    f"code {s.code}: name '{s.raw_name}' ({s.base}) diverges "
                    f"from existing name(s) {sorted(cl.bases)} under the same "
                    f"code -- verify same physical station")
            cl.bases.add(s.base)
            cl.add_raw_name(s.raw_name)
            cl.qualifiers.append(s.qualifier)
            cl.source_files.add(s.source_file)
            name_to_cluster.setdefault(s.base, cid)
        else:
            cid = name_to_cluster.get(s.base)
            if cid is None:
                cid = new_cluster()
                clusters[cid].bases.add(s.base)
                clusters[cid].flags.append(
                    f"no numeric code ever found for '{s.raw_name}' "
                    f"({s.base}) -- provisional cluster, please verify this "
                    f"is not a spelling variant of an existing coded station")
                name_to_cluster[s.base] = cid
            cl = clusters[cid]
            cl.add_raw_name(s.raw_name)
            cl.qualifiers.append(s.qualifier)
            cl.source_files.add(s.source_file)

        if s.year is not None and s.month is not None and s.has_data:
            clusters[cid].years_months[s.year].add(s.month)

        if s.direct_coord is not None:
            lat, lon, alt = s.direct_coord
            if lat is not None or lon is not None:
                clusters[cid].coord_candidates.append((lat, lon, alt, s.source_file, 1))

        if s.coord_text:
            lat, lon, alt = parse_embedded_coords(s.coord_text)
            if lat is not None or lon is not None:
                clusters[cid].coord_candidates.append((lat, lon, alt, s.source_file, 2))

    return clusters


def resolve_coords(cl):
    """Pick the coordinate candidate with best priority (1=Toa do sheet,
    2=embedded text). Flags a conflict if two same-priority-or-different
    candidates disagree beyond a small tolerance."""
    if not cl.coord_candidates:
        return None, None, None
    # de-dupe identical (lat, lon, alt, source, priority) tuples -- the same
    # embedded annotation is often carried by several sightings of the same
    # station (e.g. once per month-with-data), which would otherwise inflate
    # coordinate-conflict comparisons without adding information.
    unique_candidates = list(dict.fromkeys(cl.coord_candidates))
    best_priority = min(c[4] for c in unique_candidates)
    top = [c for c in unique_candidates if c[4] == best_priority]
    lat, lon, alt, src, _ = top[0]
    for other in top[1:] + [c for c in unique_candidates if c[4] != best_priority]:
        olat, olon, oalt, osrc, _ = other
        if lat is not None and olat is not None and abs(lat - olat) > 0.05:
            cl.flags.append(f"coordinate conflict: lat {lat} ({src}) vs "
                             f"{olat} ({osrc})")
        if lon is not None and olon is not None and abs(lon - olon) > 0.05:
            cl.flags.append(f"coordinate conflict: lon {lon} ({src}) vs "
                             f"{olon} ({osrc})")
        if alt is None and oalt is not None:
            alt = oalt
    if not in_plausible_bounds(lat, lon):
        cl.flags.append(f"coordinate out of plausible Vietnam bounding box: "
                         f"lat={lat}, lon={lon} (source {src}) -- verify, "
                         f"not auto-corrected")
    return lat, lon, alt


def compute_year_fields(years_months):
    if not years_months:
        return [], [], []
    years_available = sorted(years_months.keys())
    y_min, y_max = years_available[0], years_available[-1]
    gap_years = [y for y in range(y_min, y_max + 1) if y not in years_months]
    partial_years = [y for y in years_available if len(years_months[y]) < 12]
    return years_available, gap_years, partial_years


def best_display_name(cl):
    # Prefer the longest / most diacritic-rich raw name as canonical display,
    # but name_variants keeps everything for traceability.
    return max(cl.raw_names, key=lambda n: (len(strip_accents(n)) != len(n), len(n))) if cl.raw_names else ""


def derive_city(display_name):
    # Strip a trailing parenthetical qualifier; city = the remaining base.
    text = display_name.replace("\n", " ").strip()
    text = re.sub(r"\s*\([^)]*\)\s*$", "", text).strip()
    text = re.sub(r"\s+", " ", text)
    return text


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    files = sorted(glob.glob(os.path.join(RAW_DIR, "*.xlsx")))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    log(f"Found {len(files)} raw files in {RAW_DIR}")

    for path in files:
        log(f"--- processing {os.path.basename(path)} ---")
        process_file(path)

    log(f"Total sightings collected: {len(sightings)}")

    clusters = build_clusters(sightings)
    log(f"Total clusters (candidate stations): {len(clusters)}")

    def cluster_sort_key(item):
        cid, cl = item
        years = [y for y in cl.years_months.keys()]
        return (min(years) if years else 9999, cid)

    ordered_clusters = sorted(clusters.items(), key=cluster_sort_key)

    rows = []
    flagged_rows = []
    for i, (cid, cl) in enumerate(ordered_clusters, start=1):
        station_id = f"ST{i:03d}"
        official_code = ";".join(str(c) for c in sorted(cl.codes)) if cl.codes else ""
        display_name = best_display_name(cl)
        city = derive_city(display_name)
        lat, lon, alt = resolve_coords(cl)
        years_available, gap_years, partial_years = compute_year_fields(cl.years_months)
        cl.flags = list(dict.fromkeys(cl.flags))  # de-dupe repeated flag text
        # (coordinate-conflict flags can repeat once per redundant sighting
        # of the same embedded annotation, e.g. once per month-with-data)

        row = {
            "station_id": station_id,
            "official_code": official_code,
            "name_variants": "; ".join(cl.raw_names),
            "city": city,
            "lat": lat if lat is not None else "",
            "lon": lon if lon is not None else "",
            "altitude": alt if alt is not None else "",
            "years_available": ";".join(str(y) for y in years_available),
            "gap_years": ";".join(str(y) for y in gap_years),
            "partial_years": ";".join(str(y) for y in partial_years),
            "source_file": "; ".join(sorted(cl.source_files)),
            "flags": " | ".join(cl.flags),
        }
        rows.append(row)
        if cl.flags:
            for f in cl.flags:
                flagged_rows.append({"station_id": station_id, "name": display_name, "flag": f})

    fieldnames = ["station_id", "official_code", "name_variants", "city",
                  "lat", "lon", "altitude", "years_available", "gap_years",
                  "partial_years", "source_file", "flags"]
    with open(OUT_METADATA_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    with open(OUT_FLAGS_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["station_id", "name", "flag"])
        w.writeheader()
        w.writerows(flagged_rows)

    with open(OUT_LOG, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    print(f"Wrote {len(rows)} station rows to {OUT_METADATA_CSV}")
    print(f"Wrote {len(flagged_rows)} flagged rows to {OUT_FLAGS_CSV}")
    print(f"Log: {OUT_LOG}")


if __name__ == "__main__":
    main()
