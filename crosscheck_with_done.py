#!/usr/bin/env python3
"""
crosscheck_with_done.py

Secondary, review-only pass: cross-checks the coordinates in metadata_csv
(built by build_station_inventory.py) against the confirmed-coordinate
reference files in Excel_test/Done/, by station name.

Per project convention this NEVER writes into metadata_csv or fills a
blank lat/lon/altitude -- it only produces a companion
coordinate_crosscheck.csv listing, for every name-match found, the
metadata_csv value side-by-side with the Done-file value and whether they
agree, so a human can decide what to do with each row.

Coverage note: the 16 files in Excel_test/Done use at least two distinct
coordinate-annotation styles:
  1. embedded free text directly below the station name, e.g.
     "lg: 102 54' lt: 25 7' alt: 1893 m." (11 of 16 files) -- parsed here.
  2. dedicated lat./alt. columns in a per-station header block (the
     1949-1953 files) -- NOT parsed by this script; those files are
     listed in the report as "not covered" rather than silently treated
     as a clean bill of health.
"""

import csv
import glob
import os
import re
import sys
import unicodedata

import openpyxl

DONE_DIR = os.environ.get("DONE_DIR", "Excel_test/Done")
METADATA_CSV = os.environ.get("METADATA_CSV", "metadata_csv")
OUT_CSV = os.environ.get("OUT_CROSSCHECK_CSV", "coordinate_crosscheck.csv")

COORD_LINE_RE = re.compile(
    r"(?:lg|long|long\s*e)\.?\s*:?\s*(\d+)\s*[°o]?\s*(\d+)?['’]?.*?"
    r"(?:lt|lat|lat\s*n)\.?\s*:?\s*(\d+)\s*[°o]?\s*(\d+)?['’]?.*?"
    r"alt\.?\s*:?\s*([\d\.,]+)\s*m",
    re.IGNORECASE,
)


def strip_accents(s):
    if s is None:
        return ""
    text = str(s).replace("đ", "d").replace("Đ", "D")
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def clean_token(s):
    s = strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def dms_to_decimal(deg, minu):
    deg = float(deg)
    minu = float(minu) if minu not in (None, "") else 0.0
    return round(deg + minu / 60.0, 4)


def scan_done_files(done_dir):
    """Returns (registry, covered_files, uncovered_files).
    registry: normalized_name -> list of (lat, lon, alt, file, sheet)"""
    registry = {}
    covered, uncovered = [], []
    files = sorted(glob.glob(os.path.join(done_dir, "*.xlsx")))
    for path in files:
        fname = os.path.basename(path)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            uncovered.append(f"{fname} (open error: {e})")
            continue
        file_hits = 0
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str):
                        continue
                    m = COORD_LINE_RE.search(v)
                    if not m:
                        continue
                    lon = dms_to_decimal(m.group(1), m.group(2))
                    lat = dms_to_decimal(m.group(3), m.group(4))
                    try:
                        alt = float(m.group(5).replace(",", "."))
                    except ValueError:
                        alt = None
                    # station name: same column, one row above
                    name_cell = ws.cell(row=cell.row - 1, column=cell.column)
                    name = name_cell.value
                    if not name or not isinstance(name, str):
                        continue
                    key = clean_token(name)
                    if not key:
                        continue
                    registry.setdefault(key, []).append(
                        (lat, lon, alt, fname, sn, name.strip()))
                    file_hits += 1
        if file_hits:
            covered.append(fname)
        else:
            uncovered.append(fname)
    return registry, covered, uncovered


def main():
    registry, covered, uncovered = scan_done_files(DONE_DIR)
    print(f"Done-folder coordinate references found: {sum(len(v) for v in registry.values())} "
          f"across {len(covered)} file(s): {covered}")
    print(f"Not covered by this cross-check parser ({len(uncovered)} file(s)): {uncovered}")

    with open(METADATA_CSV, encoding="utf-8-sig") as f:
        meta_rows = list(csv.DictReader(f))

    out_rows = []
    for row in meta_rows:
        # try matching on every raw name variant, not just the display name,
        # since Done-file spellings may match an older variant better
        variants = [v.strip() for v in row["name_variants"].split(";") if v.strip()]
        variants.append(row["city"])
        matched = []
        seen_keys = set()
        for v in variants:
            key = clean_token(v)
            if not key or key in seen_keys:
                continue
            seen_keys.add(key)
            if key in registry:
                matched.extend(registry[key])

        if not matched:
            continue

        m_lat, m_lon = row["lat"], row["lon"]
        m_lat_f = float(m_lat) if m_lat not in ("", None) else None
        m_lon_f = float(m_lon) if m_lon not in ("", None) else None

        for (d_lat, d_lon, d_alt, dfile, dsheet, dname) in matched:
            lat_diff = abs(m_lat_f - d_lat) if m_lat_f is not None else None
            lon_diff = abs(m_lon_f - d_lon) if m_lon_f is not None else None
            agree = (lat_diff is not None and lat_diff <= 0.05 and
                      lon_diff is not None and lon_diff <= 0.05)
            out_rows.append({
                "station_id": row["station_id"],
                "metadata_city": row["city"],
                "metadata_lat": m_lat,
                "metadata_lon": m_lon,
                "metadata_altitude": row["altitude"],
                "done_name": dname,
                "done_lat": d_lat,
                "done_lon": d_lon,
                "done_altitude": d_alt,
                "done_source_file": dfile,
                "done_sheet": dsheet,
                "agree_within_0.05deg": agree,
                "note": ("metadata_csv lat/lon blank -- Done value shown for "
                         "your review, NOT auto-filled" if m_lat_f is None else ""),
            })

    fieldnames = ["station_id", "metadata_city", "metadata_lat", "metadata_lon",
                  "metadata_altitude", "done_name", "done_lat", "done_lon",
                  "done_altitude", "done_source_file", "done_sheet",
                  "agree_within_0.05deg", "note"]
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(out_rows)

    n_mismatch = sum(1 for r in out_rows if not r["agree_within_0.05deg"])
    print(f"Wrote {len(out_rows)} cross-check row(s) to {OUT_CSV} "
          f"({n_mismatch} disagree or metadata_csv had no coordinate to compare)")


if __name__ == "__main__":
    main()
